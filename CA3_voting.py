"""Code to implement voting rules, where voting rule is a function that takes as input 
the preferences of a set of agents and outputs a winning alternative, by loading values 
from an excel sheet. Voting rules implemented include Dictatorship, Plurality, Veto, 
Borda, Harmonic, Single Transferable Vote and RangeVoting.
"""


import copy


def generatePreferences(values):
    """A function to sort the order of alternatives for each agent, based on the agent's
    preference values, after loading an openpyxl worksheet object, and return a dictionary
    containing the agents as keys and the list or ordered alternatives as values.

    Args:
        values : An openpyxl worksheet object

    Returns:
        dictionary: A dictionary containing the agents as keys and a list of alternatives,
        in descending order of preference for each agent, as values
    """

    # Create a dictionary for storing agents as keys and list of alteratives as values
    preferences_dict = {}

    # Looping through each row in sheet
    for row in range(1, values.max_row+1):
        # As agent number is same as row number in sheet
        agent = row
        # Initialise an empty dictionary as value for each agent, 
        # for storing the alternative and the agent's corresponding preference
        preferences_dict[agent] = {}
        # Looping through the entire column values for each row
        for column in range(1, values.max_column+1):
            # As alternative number is same as column number in sheet
            alternative = column
            cell_value = values.cell(row, column).value
            # Adding each alternative and the column value as a key pair to each agent's value
            preferences_dict[agent][alternative] = cell_value
            # preferences_dict[agent].update({alternative: cell_value})

        # A dictionary storing each alternative's number and the current agent's preference
        alternatives_dict = preferences_dict[agent]

        # Getting the preference order of alternatives, by sorting the dictionary keys(alternatives), 
        # based on it's values(preferences) for each agent
        alternatives_sorted_list = sorted(alternatives_dict, key=alternatives_dict.get)[::-1]
        # Updating preferences dictionary with the list of alternatives ordered 
        # based on the agent's preferences
        preferences_dict[agent] = alternatives_sorted_list

    return preferences_dict


def dictatorship(preferenceProfile, agent):
    """A function which takes a dictionary, preferenceProfile(containing each agents 
    preferred alternatives) as input, and for a given agent returns a winner, 
    the alternative this agent ranks first.

    Args:
        preferenceProfile (dictionary): A dictionary with agents as keys and a list of 
        alternatives, in descending order of preference for each agent, as values
        agent (integer): A number representing the agent

    Returns:
        integer: Winning Alternative
    """

    try:
        # Checking if the given number corresponds to a valid agent
        if agent not in preferenceProfile.keys():
            raise ValueError
        return preferenceProfile[agent][0]
    except ValueError:
        print('{} does not correspond to an agent'.format(agent))
        # return ('{} does not correspond to an agent').format(agent)
        

def get_max_list(temp_dict):
    """A function which takes a dictionary as input, and finds the keys that has 
    the corresponding max values, and returns these keys as a list

    Args:
        temp_dict (dictionary): A dictionary containing all the alternatives as keys and 
        their scores as values

    Returns:
        list: List of possible winners
    """
    possible_winners = []
    max_value = max(temp_dict.values())
    for key, value in temp_dict.items():
        if value == max_value:
            possible_winners.append(key)
    return possible_winners     


def get_min_list(temp_dict):
    """A function which takes a dictionary as input, and finds the keys that has 
    the corresponding minimum values, and returns these keys as a list

    Args:
        temp_dict (dictionary): A dictionary containing all the alternatives as keys and 
        their frequencies as values

    Returns:
        list: List of least frequent alternatives
    """
    least_frequency = []
    min_value = min(temp_dict.values())
    for key, value in temp_dict.items():
        if value == min_value:
            least_frequency.append(key)
    return least_frequency  


def tiebreak_output(winners_list, tieBreak, preferences):
    """A function which finds the winner based on the tiebreak rule

    Args:
        list (list): A list of possible winning alternatives with same high scores
        tieBreak (integer or string): An integer or a string based on tiebreak option
        preferences (dictionary): A dictionary with agents as keys and a list of 
        alternatives, in descending order of preference for each agent, as values

    Returns:
        integer: Winning Alternative
    """
    try:
        if tieBreak == 'max':
            return max(winners_list)
        elif tieBreak == 'min':
            return min(winners_list)
        else:
            agent = tieBreak
            # Checking if the given number corresponds to a valid agent
            if agent not in preferences.keys():
                raise ValueError
            temp_dict = {}
            for alternative in winners_list:
                # For every alternative in the list of possible winners, get their corresponding
                # indexes from the agent's preferences, and store as key pair in temp_dict
                temp_dict[alternative] = preferences[agent].index(alternative)

            # Select the one that agent ranks the highest in his/her preference ordering, 
            # by selecting the alternative (key) with least value in temp_dict (least index)
            return min(temp_dict, key=temp_dict.get)
    except ValueError:
        print('The value given in tiebreak, {} does not correspond to an agent'.format(agent))


def scoringRule(preferences, scoreVector, tieBreak):
    """This function assigns a score to each of the alternative, based on their preference ordering
    in the preference profile passed as input. The score alloted for each position of alternative,
    is derived from a score vector. For every agent, the function assigns the highest score in the 
    scoring vector to the most preferred alternative of the agent, the second highest score to the 
    second most preferred alternative of the agent and so on, and the lowest score to the least 
    preferred alternative of the agent. In the end, it returns the alternative with the highest 
    total score, using the tie-breaking option to distinguish between alternatives with the same score.

    Args:
        preferences (dictionary): A dictionary containing the agents as keys and a list 
        of alternatives, in descending order of preference for each agent, as values
        scoreVector (list): A list of scores to be assigned to the alternatives
        tieBreak (integer or string): An integer or a string based on tiebreak option

    Returns:
        integer: Winning Alternative
    """
 
    # Check if sortVector length is same as the number of alternatives, else raise an error
    try:
        if len(scoreVector) != len(preferences[1]):
                raise ValueError

        # sort score vector in descending order
        scoreVector.sort(reverse=True)

        # Create a temporary dictionary to store the score for each alternative
        scoring_dict = {}
        for alternatives_list in preferences.values():
            for index, alternative in enumerate(alternatives_list):
                # update the score of the alternative in scoring_dict, based on the index postion of alternative
                # as it gets assigned the score value of the same index from the sorted scoreVector
                if alternative not in scoring_dict:
                    scoring_dict[alternative] = scoreVector[index]
                else:
                    scoring_dict[alternative] += scoreVector[index]

        possible_winners = get_max_list(scoring_dict)
        return tiebreak_output(possible_winners, tieBreak, preferences)
    except ValueError:
        print("Incorrect input")
        return False


def plurality(preferences, tieBreak):
    """A function which takes a dictionary, preferences(containing each agents 
    preferred alternatives) as input, and chooses an alternative as winner, that 
    appears the most times in the first position of the agents' preference orderings. 
    In the case of a tie, use a tie-breaking rule to select a single winner.

    Args:
        preferences (dictionary): A dictionary containing the agents as keys and a list of 
        alternatives, in descending order of preference for each agent, as values
        tieBreak (integer or string): An integer or a string based on tiebreak option

    Returns:
        integer: Winning Alternative
    """
    # Create a temporary dictionary to store the first position count of each alternative
    first_position_count = {}

    # Looping through the values of preferences, to get each agents, list of alternatives
    for alternatives_list in preferences.values():
        # get the most preferred alternative from the agents list of alternatives
        alternative = alternatives_list[0]
        # initialise the temporary dictionary for each new alternative, with count 1
        if alternative not in first_position_count:
            first_position_count[alternative] = 1
        else:
            # update the count of alternatives for each repeated occurence
            first_position_count[alternative] += 1

    # get the list of possible winners, with same high scores
    possible_winners = get_max_list(first_position_count)
    # find rhe winner using the tiebreak rule
    return tiebreak_output(possible_winners, tieBreak, preferences)
    

def veto(preferences, tieBreak):
    """A function which takes a dictionary, preferences(containing each agents 
    preferred alternatives) as input, and chooses an alternative as winner based 
    on the veto rule. Here every agent assigns 0 points to the alternative that 
    they rank in the last place of their preference orderings, and 1 point to every 
    other alternative. The winner is the alternative with the most number of points.
    In the case of a tie, use a tie-breaking rule to select a single winner.

    Args:
        preferences (dictionary): A dictionary containing the agents as keys and a list of 
        alternatives, in descending order of preference for each agent, as values
        tieBreak (integer or string): An integer or a string based on tiebreak option

    Returns:
        integer: Winning Alternative
    """
    # initialise preferences_veto with keys as alternatives, and with values as 0
    preferences_veto = dict.fromkeys(preferences[1], 0)

    # loop through each list of alternatives, and add points based on veto rule
    for alternatives_list in preferences.values():
        # Add one to the value for every 'alternative', except for the last one
        for alternative in alternatives_list[:-1]:
            preferences_veto[alternative] += 1

    # get list of possible winners (with most points)
    possible_winners = get_max_list(preferences_veto)
    # return the winner based on the tieBreak rule
    return tiebreak_output(possible_winners, tieBreak, preferences)
    

def borda(preferences, tieBreak):
    """A function which takes a dictionary, preferences(containing each agents 
    preferred alternatives) as input, and chooses an alternative as winner based 
    on the borda rule. Here every agent assigns a score of m - j to the alternative
    ranked at position j (m is the total number of alternatives). The winner is the 
    alternative with the highest score. In the case of a tie, use a tie-breaking rule 
    to select a single winner.

    Args:
        preferences (dictionary): A dictionary containing the agents as keys and a list of 
        alternatives, in descending order of preference for each agent, as values
        tieBreak (integer or string): An integer or a string based on tiebreak option

    Returns:
        integer: Winning Alternative
    """
    # initialise preferences_borda with keys as alternatives, and with values as 0
    preferences_borda = dict.fromkeys(preferences[1], 0)
    # get the number of alternatives
    num_alternatives = len(preferences[1])

    for alternatives_list in preferences.values():
        # Add value for every preference item, based on their rank position
        for rank_position, alternative in enumerate(alternatives_list, start=1):
            preferences_borda[alternative] += (num_alternatives - rank_position)

    possible_winners = get_max_list(preferences_borda)
    return tiebreak_output(possible_winners, tieBreak, preferences)
    

def harmonic(preferences, tieBreak):
    """A function which takes a dictionary, preferences(containing each agents 
    preferred alternatives) as input, and chooses an alternative as winner based 
    on the harmonic rule. Here every agent assigns a score of 1/j to the alternative
    ranked at position j. The winner is the alternative with the highest score. 
    In the case of a tie, use a tie-breaking rule to select a single winner.

    Args:
        preferences (dictionary): A dictionary containing the agents as keys and a list of 
        alternatives, in descending order of preference for each agent, as values
        tieBreak (integer or string): An integer or a string based on tiebreak option

    Returns:
        integer: Winning Alternative
    """
    # initialise preferences_borda with keys as alternatives, and with values as 0
    preferences_harmonic = dict.fromkeys(preferences[1], 0)

    for alternatives_list in preferences.values():
        # Add value for every preference item, based on their rank position
        for rank_position, alternative in enumerate(alternatives_list, start=1):
            preferences_harmonic[alternative] += (1/rank_position)

    possible_winners = get_max_list(preferences_harmonic)
    return tiebreak_output(possible_winners, tieBreak, preferences)


def STV(preferences, tieBreak):
    """A function which takes a dictionary, preferences(containing each agents 
    preferred alternatives) as input, and chooses an alternative as winner based 
    on the STV rule. The voting rule works in rounds. In each round, the alternatives 
    that appear the least frequently in the first position of agents' rankings are removed, 
    and the process is repeated. When the final set of alternatives is removed (one or 
    possibly more), then this last set is the set of possible winners. If there are more 
    than one, a tie-breaking rule is used to select a single winner.

    Args:
        preferences (dictionary): A dictionary containing the agents as keys and a list of 
        alternatives, in descending order of preference for each agent, as values
        tieBreak (integer or string): An integer or a string based on tiebreak option

    Returns:
        integer: Winning Alternative
    """
    # used copy.deepcopy to prevent modification to the original dictionary
    temp_preferences = copy.deepcopy(preferences)

    while True:

        # temporary dictionary for storing the frequency count of alternatives which resets to 0 
        # after each updation
        alternative_frequency = dict.fromkeys(temp_preferences[1], 0)
        # For each occurence of alternative as first element in the alternatives list, add a count of 1, 
        # to update its frequency
        for alternatives_list in temp_preferences.values():
            alternative_frequency[alternatives_list[0]] += 1

        # list of least frequency keys (the alternatives to be removed)
        alternatives_to_remove = get_min_list(alternative_frequency)

        # get the final set of alternatives to be removed, when its length matches the length of the
        # updated alternatives list in temp preferences dictionary
        if len(alternatives_to_remove) == len(temp_preferences[1]):
            return tiebreak_output(alternatives_to_remove, tieBreak, preferences)
        else:
            # remove the least frequent alternatives from the alternatives dictionary after each round
            for alternative in alternatives_to_remove:
                alternative_frequency.pop(alternative, None)

            # to update preferences values, by removing the least frequency keys(alternatives)
            for value in temp_preferences.values():
                for alternative in alternatives_to_remove:
                    value.remove(alternative)
            

def rangeVoting(values, tieBreak):
    """A function which loads an openpyxl worksheet object, and returns the alternative 
    that has the maximum sum of valuations, i.e., the maximum sum of numerical values 
    in the xlsx file, using the tie-breaking option to distinguish between possible winners.

    Args:
        values : An openpyxl worksheet object
        tieBreak (integer or string): An integer or a string based on tiebreak option

    Returns:
        integer: Winning Alternative
    """
    valuation_sum = {}
    
    # Looping through each column in sheet (corresponding to an alternative)
    for column in range(1, values.max_column+1):
        alternative = column

        # initialising the key (alternative) to the valuation_sum dictionary
        if alternative not in valuation_sum.keys():
            valuation_sum[alternative] = 0

        # Looping through the entire row values for each column
        for row in range(1, values.max_row+1):
            # Adding the cell value to the corresponding alternative, to get the total sum
            valuation_sum[alternative] += values.cell(row, column).value

    possible_winners = get_max_list(valuation_sum)
    return tiebreak_output(possible_winners, tieBreak, generatePreferences(values))


if __name__ == '__main__':
    import openpyxl
    import pprint

    workbook = openpyxl.load_workbook('/projects/voting.xlsx')
    sheet = workbook.active

    pp = pprint.PrettyPrinter(indent=4, width = 120)

    # print("Preferences Dictionary:")
    # pprint.pprint(generatePreferences(sheet))