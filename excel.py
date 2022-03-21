import openpyxl
import os

# Go to folder where excel file is located
os.chdir('/file_path') # enter 'file_path', as the path to folder where excel file is stored
print(os.getcwd())

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook('voting.xlsx')
print(wb.sheetnames)

sheet = wb['Sheet1']

# wb.create_sheet('Sheet2', index=0)
# wb.remove(sheet)

print(sheet['A1'].value)

# Initilizing the dictionary with agents as keys, based on max row in sheet
preferences_dict = {key: [] for key in range(1, sheet.max_row+1)}
# preferences_dict = {}
print('Initialized preferences dictionary:', preferences_dict)

print(sheet.max_row, sheet.max_column)
for row in range(1, sheet.max_row+1):
    # if row not in preferences_dict.keys():
    #     preferences_dict[row] = []
    for col in range(1, sheet.max_column+1):
        preferences_dict[row].append(sheet.cell(row, col).value)
        # print(sheet.cell(row, col).value)


print('Populated preferences dictionary: ', preferences_dict)
